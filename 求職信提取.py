#!/usr/bin/env python
# coding: utf-8

# In[1]:


import win32com.client
import pandas as pd
import openpyxl
import datetime
import re
import xlwt
from openpyxl import Workbook
import os.path


# In[2]:


#確認擬寫入的excel是否存在，如不存在則以年份建立一新檔案
def set_document():
    today_year=datetime.datetime.now().year
    filename='K:\\Human Resource\\05招募甄選\\04面談紀錄\\99報表\\'+str(today_year)+' HR Candidate Collection.xlsx' #依照一份先給予檔名
    if os.path.isfile(filename):  #確認檔案是否存在
        print ("File exist: "+filename)
    else:                         #如不存在則建立一檔案，並先寫入欄位名稱
        print ("File not exist")
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        writer.save()
        writer.close()
        wb = openpyxl.load_workbook(filename=filename)
        sheet = wb['Sheet1']
        new_row = ['DATE','NAME','POSITION','EMAIL','PHONE NO.','CHANNEL','View Profile/Resume']
        sheet.append(new_row)
        wb.save(filename)
        print ("Creat new file: "+filename )
    return filename


# In[3]:


#JOBSDB求職者資訊提取
def jobsdb_extract(filename):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-Jobsdb"].Folders["jobsdb-not processed"]
    inbox_proccessed = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-Jobsdb"]
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  #按照日期排序
    #SenderName,SenderEmailAddress,Subject,Body,ReceivedTime,ConversationID​
    email_count=0
    while email_count>=0:
        try:
            message = messages.Restrict("[SentOn] > '7/12/2020 04:30 PM'")[email_count]  
            data=message.Body
            split_data=re.split("\n|\r|\t",data)
            split_data = [x for x in split_data if x != '']
            split_data = [x for x in split_data if x != ' ']
            split_data
            #email incoming date
            try: 
                received_date_temp=message.ReceivedTime
                received_date=str(received_date_temp)
                received_date=received_date[0:10]
                received_date =  datetime.datetime.strptime(received_date, '%Y-%m-%d').date()
            except:
                received_date=""
            
            
            #candidate name extract
            try:
                from_match=re.search('from ',message.Subject)
                for_match=re.search('for ',message.Subject)
                candidate_name=message.Subject[from_match.end():for_match.start()]#求職者姓名
            except:
                candidate_name=""
                
                
            #candidate job extract
            try:
                for_JHK=re.search('\(JHK',message.Subject)
                candidate_job=message.Subject[for_match.end():for_JHK.start()]#求職工作
            except:
                candidate_job=""
            
            #candidate phone extract
            try:
                for line in split_data:
                    #line.replace()
                    if len(line)==8 and line[0] in ('1','2','3','4','5','6','7','8','9'):
                        phone=line
                        break
                    else:
                        phone=""
            except:
                phone=""

            #candidate email extract
            try:
                email = [s for s in split_data if "@" in s]
                if len(email[0])<=60:
                    email_ext=email[0]
                    email_ext
                else:
                    email_ext=""
            except:
                email_ext=""

            #resume url extract
            try:
                resume_url = [s for s in split_data if "Download resume" in s]
                resume_link=resume_url[0].replace("Download resume <","")
                resume_link=resume_link.replace(">","")
            except:
                resume_link=""
                
            #split_data
            #print("姓名:",candidate_name,"\n職位:",candidate_job,"\n電話: ",phone,"\nemail: ",email_ext,"\nresume url:",resume_link)        
            
            #wirte into excel 
            wb = openpyxl.load_workbook(filename=filename)
            sheet = wb['Sheet1']
            new_row = [received_date,candidate_name, candidate_job,email_ext, phone,'JOBSDB',resume_link]
            sheet.append(new_row)
            wb.save(filename)
            
            email_count+=1
            
        except: 
            if email_count==0 :
                print('jobsDB無求職信件，無須提取')
            else:
                print('已完成jobsDB求職者進件日期、姓名、職位、email、電話及履歷資訊提取及寫入，共'+str(email_count)+'筆')
            return 


# In[4]:


#JOBSDB將已提取email放置已處理區
def jobsdb_move():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-Jobsdb"].Folders["jobsdb-not processed"]
    inbox_proccessed = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-Jobsdb"]
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  #按照日期排序
    #SenderName,SenderEmailAddress,Subject,Body,ReceivedTime,ConversationID​
    email_count=0
    move_count=0
    while email_count==0:
        try:
            message = messages.Restrict("[SentOn] > '7/12/2020 04:30 PM'")[email_count]  
            #move to 履歷已處理區
            message.move(inbox_proccessed)
            move_count+=1
        except:
            if move_count==0:
                print("jobsDB無信件需移動")
            else:
                print("jobsDB已搬至履歷已處理區，共"+str(move_count)+"筆")
            return


# In[5]:


#CTgoodjobs求職者資訊提取
def CTgoodjobs_extract(filename):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-CTgoodjobs"].Folders["CTgoodjobs-not processed"]
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  #按照日期排序
    #SenderName,SenderEmailAddress,Subject,Body,ReceivedTime,ConversationID​
    email_count=0
    while email_count>=0:
        try:
            message = messages.Restrict("[SentOn] > '7/12/2021 04:30 PM'")[email_count]  
            data=message.Body
            split_data=re.split("\n|\r|\t",data)
            split_data = [x for x in split_data if x != '']
            split_data = [x for x in split_data if x != ' ']
            #split_data

            #email incoming date
            try:
                received_date_temp=message.ReceivedTime
                received_date=str(received_date_temp)
                received_date=received_date[0:10]
                received_date =  datetime.datetime.strptime(received_date, '%Y-%m-%d').date() 
            except:
                received_date=""
            
            #candidate name extract
            try:
                candidate_name_temp = [s for s in split_data if "Name: " in s]
                candidate_name=candidate_name_temp[0].replace("Name: ","")
            except:
                candidate_name=""

            #candidate job extract
            try:
                candidate_job_temp = [s for s in split_data if "Application for the position of " in s]
                candidate_job=candidate_job_temp[0].replace("Application for the position of ","")
            except:
                candidate_job=""
            
            #candidate phone extract
            try:
                phone_temp = [s for s in split_data if "Contact No.: " in s]
                phone=phone_temp[0].replace("Contact No.:","") 
            except:
                phone=""

            #candidate email extract
            try:
                email_temp = [s for s in split_data if "E-mail: " in s]
                email=email_temp[0].replace("E-mail: ","")
            except:
                email=""
            
            #resume url extract
            try:
                resume_url_temp = [s for s in split_data if "View Resume " in s]
                resume_url=resume_url_temp[0].replace("View Resume <","")
                resume_url=resume_url.replace(" ","")
                resume_url=resume_url.replace(">","")
            except:
                resume_url=""
            
            #wirte into excel 
            wb = openpyxl.load_workbook(filename=filename)
            sheet = wb['Sheet1']
            new_row = [received_date,candidate_name, candidate_job,email, phone,'CTgoodjobs',resume_url]
            sheet.append(new_row)
            wb.save(filename)
            email_count+=1
        except: 
            if email_count==0:
                print('CTgoodjobs無求職信件，無須提取')
            else:
                print('已完成CTgoodjobs求職者進件日期、姓名、職位、email、電話及履歷資訊提取及寫入，共'+str(email_count)+'筆')
            return 

    #print("姓名:",candidate_name,"\n職位:",candidate_job,"\n電話: ",phone,"\nemail: ",email,"\nresume url:",resume_url)
    #split_data


# In[6]:


#CTgoodjobs將已提取email放置已處理區
def CTgoodjobs_move():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-CTgoodjobs"].Folders["CTgoodjobs-not processed"]
    inbox_proccessed = outlook.Folders["HR9018@email.esunbank.com.tw"].Folders["收件匣"].Folders["應聘-CTgoodjobs"]
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  #按照日期排序
    #SenderName,SenderEmailAddress,Subject,Body,ReceivedTime,ConversationID​
    email_count=0
    move_count=0
    while email_count==0:
        try:
            message = messages.Restrict("[SentOn] > '7/12/2020 04:30 PM'")[email_count]  
            #move to 履歷已處理區
            message.move(inbox_proccessed)
            move_count+=1
        except:
            if move_count==0:
                print("CTgoodjobs無信件需移動")
            else:
                print("CTgoodjobs已搬至履歷已處理區，共"+str(move_count)+"筆")
            return


# In[7]:


#主函式
def main():
    starttime=datetime.datetime.now()
    filename=set_document()
    jobsdb_extract(filename)
    jobsdb_move()
    CTgoodjobs_extract(filename)
    CTgoodjobs_move()
    endtime=datetime.datetime.now()
    print('共花費：'+str(endtime-starttime)+' 秒')


# In[8]:


main()






